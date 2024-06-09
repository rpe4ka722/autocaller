from django.shortcuts import render, redirect
import requests
import os
import time
import configparser
from main.tasks import start_caller
from main.forms import AbonentForm
from main.scripts import parse_ami_response
from main.models import Abonent, SoundFile, CallList, Report, Call
from django.shortcuts import get_object_or_404
from django.conf import settings
import subprocess
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font




@login_required(login_url='account:login')
def index(request):
    config = configparser.ConfigParser()
    config.read('./django-files/config.ini')
    dep = request.user.department
    lists = CallList.objects.filter(department=dep)
    try:
        rep = Report.objects.get(department=dep, in_progress=True)
        call_list = rep.list.abonents_count() 
        percents = int(100*(rep.unchecked_abonents + rep.checked_abonents)/call_list)
        print(percents)
        context = {'report': rep,'calls_count': call_list, 'abonent_confirmed': rep.checked_abonents, 
                   'abonent_unconfirmed': rep.unchecked_abonents, 'percents': percents}
        return render(request, 'main/templates/index_in_progress.html', context)   
    except Report.DoesNotExist:
        try:
            s = requests.Session() 
            response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] + "/asterisk/rawman?action=login&username=" 
                             + config['asterisk']['username'] + "&secret=" + config['asterisk']['secret'], timeout=0.1)
            response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] +  "/asterisk/rawman?action=PJSIPShowRegistrationsOutbound", timeout=0.1)
            if parse_ami_response(response.text)['Registered'] == '1':
                status = True
            else: 
                status = False
            response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] + "/asterisk/rawman?action=logoff", timeout=0.1)    
        except:
            status = False
        context = {'lists': lists, 'status': status}
        return render(request, 'main/templates/index.html', context)


@login_required(login_url='account:login')       
def percent_status(request, report_id):
    try:
        report = Report.objects.get(id=report_id)
        abonents_count = report.list.abonents.all().count()
        percents = int(100*(report.unchecked_abonents + report.checked_abonents)/abonents_count)
        print(percents)
        if report.in_progress == True:    
            context = {'percents': percents, 'calls_count': abonents_count, 'in_progress': True, 
                       'abonent_confirmed': report.checked_abonents, 'abonent_unconfirmed': report.unchecked_abonents}
            return  JsonResponse(context)
        else:
            context = {'percents': percents, 'calls_count': abonents_count, 'in_progress': False, 
                       'abonent_confirmed': report.checked_abonents, 'abonent_unconfirmed': report.unchecked_abonents}
            return  JsonResponse(context)
    except Report.DoesNotExist:
        return HttpResponse('Неизвестная ошибка', status=500)    


@login_required(login_url='account:login')
def ping(request):
    try:
       s = requests.Session() 
       config = configparser.ConfigParser()
       config.read('./django-files/config.ini')
       response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] + "/asterisk/rawman?action=login&username=" 
                         + config['asterisk']['username'] + "&secret=" + config['asterisk']['secret'], timeout=0.1)
       response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] +  "/asterisk/rawman?action=PJSIPShowRegistrationsOutbound", timeout=0.1)
       #print(response.text)
       if parse_ami_response(response.text)['Registered'] == '1':
           status = True
       else: 
           status = False
       response = s.get("http://" + config['asterisk']['host'] + ":" + config['asterisk']['http_port'] + "/asterisk/rawman?action=logoff", timeout=0.1)    
    except:
       status = False
    context = {'status': status}
    return render(request, 'main/templates/htmx_ping.html', context)


@login_required(login_url='account:login')
def abonents(request, msg=''):
    dep = request.user.department
    adm = False
    if dep == 'ALL':
        objects = Abonent.objects.all()
        adm = True
    else:
        objects = Abonent.objects.filter(department=dep)
    context = {'objects': objects, 'msg': msg, 'adm': adm}
    return render(request, 'main/templates/abonent.html', context)


@login_required(login_url='account:login')
def create_abonent(request):
    dep = request.user.department
    if request.method == 'POST':
        form = AbonentForm(request.POST or None)
        if form.is_valid():
            msg = ''
            cd = form.cleaned_data
            mobile_phone_number = '+7' + cd['mobile_phone_number']
            if cd['secondary_mobile_phone_number'] is not None:
                secondary_mobile_phone_number = '+7' + cd['secondary_mobile_phone_number']
            else:
                secondary_mobile_phone_number = None
            phone_list = Abonent.objects.all().values_list('mobile_phone_number', flat=True)
            sec_phone_list = Abonent.objects.all().values_list('secondary_mobile_phone_number', flat=True).exclude(secondary_mobile_phone_number=None)
            if mobile_phone_number in phone_list:
                abon = Abonent.objects.get(mobile_phone_number=mobile_phone_number)
                num = cd['mobile_phone_number']
                msg = f'Ошибка! Номер { num } указан как мобильный для абонента {abon.full_name()}.'
            elif mobile_phone_number in sec_phone_list:
                num = cd['mobile_phone_number']
                abon = Abonent.objects.get(secondary_mobile_phone_number=mobile_phone_number)
                msg = f'Ошибка! Номер { num } указан как дополнительный для абонента {abon.full_name()}.'
            elif secondary_mobile_phone_number in phone_list:
                abon = Abonent.objects.get(mobile_phone_number=secondary_mobile_phone_number)
                num = cd['secondary_mobile_phone_number']
                msg = f'Ошибка! Номер { num } указан как мобильный для абонента {abon.full_name()}.'
            elif secondary_mobile_phone_number in sec_phone_list:
                abon = Abonent.objects.get(secondary_mobile_phone_number=secondary_mobile_phone_number)
                num = cd['secondary_mobile_phone_number']
                msg = f'Ошибка! Номер { num } указан как дополнительный для абонента {abon.full_name()}.'
            else:
                abon = Abonent(first_name=cd['first_name'], last_name=cd['last_name'], patronymic=cd['patronymic'], 
                               work_phone_number=cd['work_phone_number'], mobile_phone_number=mobile_phone_number, 
                               secondary_mobile_phone_number=secondary_mobile_phone_number, department=dep)
                abon.save()
        if msg != '':
            return HttpResponse(msg, status=500)
    print('End')
    return HttpResponse(msg, status=200)



@login_required(login_url='account:login')
def delete_abonent(request, id):
    try:
        object = get_object_or_404(Abonent, pk=id)
        object.delete()
    except Abonent.DoesNotExist:
        pass
    return redirect('main:abonents')


@login_required(login_url='account:login')
def edit_abonent(request, id):
    msg = ''
    try:
        object = get_object_or_404(Abonent, pk=id)    
        if request.method == 'POST':
            form = AbonentForm(request.POST or None)
            if form.is_valid():
                cd = form.cleaned_data
                mobile_phone_number = '+7' + cd['mobile_phone_number']
                if cd['secondary_mobile_phone_number'] is not None:
                    secondary_mobile_phone_number = '+7' + cd['secondary_mobile_phone_number']
                else:
                    secondary_mobile_phone_number = None
                phone_list = Abonent.objects.all().exclude(id=id)
                phone_list = phone_list.values_list('mobile_phone_number', flat=True)
                sec_phone_list = Abonent.objects.all().exclude(id=id)
                sec_phone_list = sec_phone_list.values_list('secondary_mobile_phone_number', flat=True).exclude(secondary_mobile_phone_number=None)
                if mobile_phone_number in phone_list:
                    abon = Abonent.objects.get(mobile_phone_number=mobile_phone_number)
                    num = cd['mobile_phone_number']
                    msg = f'Ошибка! Номер { num } указан как мобильный для абонента {abon.full_name()}.'
                elif mobile_phone_number in sec_phone_list:
                    num = cd['mobile_phone_number']
                    abon = Abonent.objects.get(secondary_mobile_phone_number=mobile_phone_number)
                    msg = f'Ошибка! Номер { num } указан как дополнительный для абонента {abon.full_name()}.'
                elif secondary_mobile_phone_number in phone_list:
                    abon = Abonent.objects.get(mobile_phone_number=secondary_mobile_phone_number)
                    num = cd['secondary_mobile_phone_number']
                    msg = f'Ошибка! Номер { num } указан как мобильный для абонента {abon.full_name()}.'
                elif secondary_mobile_phone_number in sec_phone_list:
                    abon = Abonent.objects.get(secondary_mobile_phone_number=secondary_mobile_phone_number)
                    num = cd['secondary_mobile_phone_number']
                    msg = f'Ошибка! Номер { num } указан как дополнительный для абонента {abon.full_name()}.'
                else:
                    object.first_name = cd['first_name']
                    object.last_name = cd['last_name']
                    object.patronymic = cd['patronymic']
                    object.work_phone_number = cd['work_phone_number']
                    object.mobile_phone_number = mobile_phone_number
                    object.secondary_mobile_phone_number = secondary_mobile_phone_number
                    object.save()
    except Abonent.DoesNotExist:
        pass
    if msg == '':
        return HttpResponse(msg, status=200)
    else: 
        return HttpResponse(msg, status=500)
    

@login_required(login_url='account:login')
def sounds(request):
    dep = request.user.department
    files = SoundFile.objects.filter(department = dep)
    time = datetime.now().timestamp
    context = {'files': files, 'time':time}
    response = render(request, 'main/templates/sounds.html', context)
    return response


@login_required(login_url='account:login')
def create_sound(request):
    dep = request.user.department
    if request.method == 'POST':
        filename = request.POST['filename'] + '.wav'
        text = request.POST['text']
        directory_temp = settings.MEDIA_ROOT + '/temp/'
        directory = settings.MEDIA_ROOT + '/sounds/' + dep + '/'
        try:
            sound_file = SoundFile.objects.get(filename=filename, department=dep)
        except SoundFile.DoesNotExist:
            sound_file = SoundFile(filename = filename, department = dep, dir = directory)
        path_temp = directory_temp + filename
        cmd = 'echo "' + text + '"| RHVoice-test -p Elena+CBL -o ' + path_temp
        cmd2 = 'sox ' + path_temp + ' --channels 1 ' + sound_file.get_full_path() + ' rate 8000'  
        try:
            result = subprocess.run([cmd], shell=True)
            if result.returncode == 0:
                try:
                   result = subprocess.run([cmd2], shell=True)
                   os.remove(path_temp)
                   print(result)
                except subprocess.CalledProcessError as e:
                    print(f"An error occurred: {e}")     
        except subprocess.CalledProcessError as e:
            print(f"An error occurred: {e}")
        sound_file.save()
    return redirect('main:sounds')


@login_required(login_url='account:login')
def download_sound(request):
    dep = request.user.department
    if request.method == 'POST':
        directory_temp = settings.MEDIA_ROOT + '/temp/'
        directory = settings.MEDIA_ROOT + '/sounds/' + dep + '/'
        file = request.FILES['sound_file']
        name = file.name
        filename = os.path.splitext(name)[0] + '.wav'
        try:
            soundfile = SoundFile.objects.get(filename=filename, department=dep)
        except SoundFile.DoesNotExist:
            soundfile = SoundFile(filename = filename, department = dep, dir = directory)
        path_temp = directory_temp + name
        cmd = 'sox ' + path_temp + ' --channels 1 ' + soundfile.get_full_path() + ' rate 8000'
        with open(path_temp, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        try:
            result = subprocess.run([cmd], shell=True)
            os.remove(path_temp)
            print(result)
        except subprocess.CalledProcessError as e:
            print(f"An error occurred: {e}")
        soundfile.save()
    return redirect('main:sounds')


@login_required(login_url='account:login')
def delete_sound(request, id):
    try:
        object = get_object_or_404(SoundFile, pk=id)
        os.remove(object.get_full_path())
        object.delete()
    except Abonent.DoesNotExist:
        pass
    return redirect('main:sounds')


@login_required(login_url='account:login')
def lists(request):
    dep = request.user.department
    lists = CallList.objects.filter(department=dep)
    sound_list = SoundFile.objects.filter(department=dep).values_list("filename", flat=True)
    abonents = Abonent.objects.filter(department=dep)
    context = {'abonents': abonents, 'lists': lists, 'sound_list': sound_list}
    response = render(request, 'main/templates/lists.html', context)
    return response


@login_required(login_url='account:login')
def create_list(request):
    dep = request.user.department
    current_user = request.user
    if request.method == 'POST':
        listname = request.POST['listname']
        list_text = request.POST['list_text']
        abonents_list = request.POST['abonents_list'].split(',')
        phones = request.POST.getlist('phone')
        try:
            accept_code = int(request.POST['accept_code'])
        except:
             return HttpResponse('Неизвестная ошибка', status=500)
        if abonents_list == '':
            return HttpResponse('Список абонентов не должен быть пустым', status=500)
        else:
            try:
                sound = SoundFile.objects.get(filename=request.POST['sound_name'])
            except SoundFile.DoesNotExist:
                return HttpResponse('Неизвестная ошибка', status=500)
            list = CallList.objects.create(list_name=listname, list_description=list_text, sound=sound, last_edit_user=current_user,
                                           accept_combination=accept_code, department=dep)
            if 'mobile' in phones:
                list.main_phone = True
            else:
                list.main_phone = False
            if 'secondary' in phones:
                list.second_phone = True
            if 'work' in phones:
                list.work_phone = True
            list.save()
            for abonent in abonents_list:
                last_name = abonent.split(' ')[0]
                first_name = abonent.split(' ')[1]
                patronymic = abonent.split(' ')[2]
                try:
                    abon = Abonent.objects.get(first_name=first_name, last_name=last_name, patronymic=patronymic)
                    list.abonents.add(abon)
                    list.save()
                except Abonent.DoesNotExist:
                    return HttpResponse('Неизвестная ошибка', status=500)
            return HttpResponse(status=200)


@login_required(login_url='account:login')
def delete_list(request, list_id):
    try:
        object = get_object_or_404(CallList, pk=list_id)
        object.delete()
    except CallList.DoesNotExist:
        pass
    return redirect('main:lists')


@login_required(login_url='account:login')
def report(request):
    dep = request.user.department
    reports = Report.objects.filter(department=dep).order_by("-start_time")
    context = {'reports': reports}
    response = render(request, 'main/templates/report.html', context)
    return response


@login_required(login_url='account:login')
def start_list(request, list_id):
    print('Вход во вью')
    if request.method == 'GET':
        user = request.user
        user_id = user.id
        report_id = start_caller.apply_async(args=[list_id, user_id], queue='hipri', routing_key='hipri')
        print(f'Старт листа вью {report_id}')
        while True:
             if report_id.ready():
                 break
             time.sleep(0.05)
        report = Report.objects.get(id=report_id.result)
        print('Выход из вью')
        response = redirect('main:index')
        return response


@login_required(login_url='account:login')
def report_delete(request, report_id):
    try:
        report = Report.objects.get(id=report_id)
        report.delete()
        response = redirect('main:report')
        return response
    except Report.DoesNotExist:
        response = redirect('main:report')
        return response


@login_required(login_url='account:login')
def report_export(request, report_id):
    try:
        report = Report.objects.get(id=report_id)
        if report.in_progress == True:
            return HttpResponse('Файл выгрузки доступен после завершения обзвона', status=500)
    except Report.DoesNotExist:
        return HttpResponse('Неизвестная ошибка', status=500)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    wb = Workbook()
    ws = wb.active
    report_dict = {'Дата проведения оповещения':report.start_time.strftime('%d.%m.%Y'), 'Начало оповещения': report.start_time.strftime('%X'), 
                   'Время завершения оповещения': report.end_time.strftime('%X'), 'Запущено пользователем': report.create_by_user.get_full_name(), 'Наименование листа оповещения': report.list.list_name,
                   'Количество абонентов': report.list.abonents_count(),'Оповещено абонентов': report.checked_abonents}
    row_num = 1
    
    font = Font(name='TimesNewRoman', sz=11, bold=True)
    font2 = Font(name='TimesNewRoman', sz=11, bold=False)
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=False,
                          indent=0)
    for name, attr in report_dict.items():
        ws.cell(row_num, 1, name)
        ws.cell(row_num, 2, attr)
        ws.cell(row_num, 1).font = font
        ws.cell(row_num, 1).border = border
        ws.cell(row_num, 1).alignment = alignment
        ws.cell(row_num, 2).font = font
        ws.cell(row_num, 2).border = border
        ws.cell(row_num, 2).alignment = alignment
        row_num += 1
    row_num += 1

    columns = ['Абонент', 'Тип номера', 'Номер телефона', 'Уведомлен','Введенный код', 'Количество неверных попыток',
               'Время начала вызова', 'Время завершения вызова', 'Ответ абонета', 'Код завершения']
    for col_num in range(len(columns)):
        ws.cell(row_num, col_num + 1, columns[col_num])
        ws.cell(row_num, col_num + 1).font = font
        ws.cell(row_num, col_num + 1).border = border
        ws.cell(row_num, col_num + 1).alignment = alignment
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30

    row_num += 1

    calls = Call.objects.filter(report=report)
    for call in calls:
        if call.confirmed:
            conf = 'Да'
        else: 
            conf = 'Нет'
        if call.call_answered:
            answ = 'Да'
        else: 
            answ = 'Нет'
        if call.call_timeout:
            end_code = 'Таймаут'
        elif call.call_error:
            end_code = 'Ошибка станции'
        elif call.ats_no_answer:
            end_code = 'Нет ответа от удаленной станции'
        elif call.asterisk_no_answer:
            end_code = 'Нет ответа от астериская'
        else: 
            end_code = call.end_code
        call_list = [call.abonent.full_name(), call.phone_type, call.abonent_number, conf, call.user_input, call.incorrect_input_count,
                     call.start_time.strftime('%X'), call.end_time.strftime('%X'), answ, end_code]
        for col_num in range(len(call_list)):
             ws.cell(row_num, col_num + 1, call_list[col_num])
             ws.cell(row_num, col_num + 1).font = font2
             ws.cell(row_num, col_num + 1).border = border
             ws.cell(row_num, col_num + 1).alignment = alignment
        row_num += 1
    wb.save(response)
    return response


def report_time_status(request, report_id):
    try:
        report = Report.objects.get(id=report_id)
        if report.in_progress == True:
            return render(request, 'main/templates/loader.html')
        else:
            context = {'report': report}
            return render(request, 'main/templates/report_time.html', context)
    except Report.DoesNotExist:
        return HttpResponse('Неизвестная ошибка', status=500)
    

def report_abon_status(request, report_id):
    try:
        report = Report.objects.get(id=report_id)
        if report.in_progress == True:
            return render(request, 'main/templates/loader.html')
        else:
            context = {'report': report}
            return render(request, 'main/templates/report_abon_count.html', context)
    except Report.DoesNotExist:
        return HttpResponse('Неизвестная ошибка', status=500)
    
